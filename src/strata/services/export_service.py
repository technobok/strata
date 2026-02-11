"""Export service for generating downloadable files from query results."""

import io
import tempfile
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate_xlsx(
    columns: list[str],
    rows: list[tuple[Any, ...]],
    sheet_name: str = "Results",
) -> bytes:
    """Generate an XLSX file from query results.

    Returns the file contents as bytes.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = sheet_name

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Create Excel table if there is data
    if rows and columns:
        last_col = get_column_letter(len(columns))
        last_row = len(rows) + 1
        table_ref = f"A1:{last_col}{last_row}"

        table = Table(displayName="Results", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in rows[:100]:  # Sample first 100 rows
            val = row[col_idx - 1]
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_xlsx_from_cache(result_hash: str, sheet_name: str = "Results") -> bytes | None:
    """Generate XLSX from a cached result."""
    from strata.services.cache_service import read_result

    columns, rows, _ = read_result(result_hash)
    if not columns:
        return None

    return generate_xlsx(columns, rows, sheet_name)


def generate_parquet_from_cache(result_hash: str) -> bytes | None:
    """Generate Parquet bytes from a cached DuckDB result."""
    from strata.services.cache_service import cache_path

    path = cache_path(result_hash)
    if not path.exists():
        return None

    conn = duckdb.connect(str(path), read_only=True)
    try:
        with tempfile.NamedTemporaryFile(suffix=".parquet") as tmp:
            conn.execute(f"COPY results TO '{tmp.name}' (FORMAT PARQUET)")
            tmp.seek(0)
            return tmp.read()
    finally:
        conn.close()


_FORMAT_MAP: dict[str, tuple[str, str]] = {
    "xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsx",
    ),
    "parquet": (
        "application/vnd.apache.parquet",
        ".parquet",
    ),
}


def generate_download(
    result_hash: str,
    fmt: str,
    sheet_name: str = "Results",
) -> tuple[bytes, str, str] | None:
    """Generate a download in the requested format.

    Returns (bytes, mimetype, extension) or None if the cache is missing.
    Raises ValueError for unsupported formats.
    """
    if fmt not in _FORMAT_MAP:
        raise ValueError(f"Unsupported format: {fmt}")

    mimetype, extension = _FORMAT_MAP[fmt]

    if fmt == "xlsx":
        data = generate_xlsx_from_cache(result_hash, sheet_name)
    elif fmt == "parquet":
        data = generate_parquet_from_cache(result_hash)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    if data is None:
        return None

    return data, mimetype, extension
